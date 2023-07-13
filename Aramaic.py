#!/usr/bin/env python
# -*- coding: utf-8  -*-

import os
import requests
import json
import hebrew_numbers
import re
from utils import *
import openpyxl
from openpyxl.cell.cell import MergedCell

from enum import Enum

class Gender(Enum):
	MALE = "זכר"
	FEMALE = "נקבה"
	M = MALE
	F = FEMALE

class Count(Enum):
	SINGULAR = "יחיד"
	PLURAL = "רבים"
	S = SINGULAR
	P = PLURAL

class Person(Enum):
	A = "ראשון"
	B = "שני"
	C = "שלישי"
	P1 = A
	P2 = B
	P3 = C

class Stem(Enum):
	PEIL = "פְּעִל(קל)"
	PAEL = "פַּעַל(פיעל)"
	ITHPEIL = "אתפְּעִל(נפעל)"
	AFEL = "אפעל(הפעיל)"
	ITPAAL = "אתפַּעַל(פועל)"
	ITTAFAL = "אתַּפְעַל(הופעל)"

class Tense(Enum):
	PAST = "עבר"
	BEINONI = "בינוני"
	BEINONI_PASSIVE = "בינוני פעול"
	PRESENT = "הווה"
	PRESENT_PASSIVE = "הווה פעול"
	FUTURE = "עתיד"
	IMPERATIVE = "ציווי"
	MAKOR = "מקור"

class Word:
	def __init__(self, text, yemenitext, translation=None):
		self.text = text
		self.yemenitext = yemenitext
		self.translation = translation

	def __repr__(self):
		return "<%s(%s)>"%(self.__class__.__name__, self.text)

	@property
	def baretext(self):
		return strip_hebrew_points(self.text)

	@property
	def htmltext(self):
		return re.sub("\[([^]+])\]", '<span class="lectionis">\\1</span>', self.text)

	@property
	def fulltext(self):
		return re.sub("\[([^]+])\]", "\\1", self.text)

	@property
	def defectivetext(self):
		return re.sub("\[[^]+]\]", "", self.text)

class Adverb(Word):
	def __init__(self, text, yemenitext, translation=None):
		super().__init__(text, yemenitext, translation)

	@property
	def morphology(self):
		return "תואר הפועל"

class InflectedAdverb(Word):
	def __init__(self, text, yemenitext, translation, person, gender, kount):
		super().__init__(text, yemenitext, translation)
		self.person = person
		self.gender = gender
		self.kount = kount

	@property
	def morphology(self):
		return ", ".join([v for v in ["תואר הפועל", "גוף %s"%self.person.value if self.person else "",
				self.gender.value if self.gender else "", self.kount.value] if v])

class PlaceName(Word):
	def __init__(self, text, yemenitext):
		super().__init__(text, yemenitext)

	@property
	def morphology(self):
		return "שם פרטי"

class PersonName(Word):
	def __init__(self, text, yemenitext, gender):
		super().__init__(text, yemenitext)
		self.gender = gender

	@property
	def morphology(self):
		return ", ".join(["שם פרטי", self.gender.value])

class Pronoun(Word):
	def __init__(self, text, yemenitext, translation, person, gender, kount):
		super().__init__(text, yemenitext, translation)
		self.person = person
		self.gender = gender
		self.kount = kount

	@property
	def morphology(self):
		return ", ".join([v for v in ["כינוי", "גוף %s"%self.person.value, self.gender.value if self.gender else "", self.kount.value] if v])

class Noun(Word):
	def __init__(self, text, yemenitext, translation, gender, kount):
		super().__init__(text, yemenitext, translation)
		self.gender = gender
		self.kount = kount

	@property
	def morphology(self):
		return ", ".join(["שם עצם", self.gender.value, self.kount.value])

class InflectedNoun(Noun):
	def __init__(self, text, yemenitext, translation, gender, kount, operson, ogender, okount):
		super().__init__(text, yemenitext, translation, gender, kount)
		self.operson = operson
		self.ogender = ogender
		self.okount = okount

	@property
	def morphology(self):
		return "שם עצם, %s, %s + גוף %s, %s%s"%(self.gender.value, self.kount.value,
				self.operson.value, "%s, "%self.ogender.value if self.ogender else "", self.okount.value)

class Numeral(Word):
	def __init__(self, text, yemenitext, translation, value, gender):
		super().__init__(text, yemenitext, translation)
		self.value = value
		self.gender = gender

class Cardinal(Numeral):
	def __init__(self, text, yemenitext, translation, value, gender):
		super().__init__(text, yemenitext, translation, value, gender)

	@property
	def morphology(self):
		return ", ".join(["מספר מונה", str(self.value), self.gender.value])

class Ordinal(Numeral):
	def __init__(self, text, yemenitext, translation, value, gender):
		super().__init__(text, yemenitext, translation, value, gender)

	@property
	def morphology(self):
		return ", ".join(["מספר סודר", str(self.value), self.gender.value])

class VerbTable:
	def __init__(self, root):
		self.root = root
		self.items = []

	def conjugate(self, stem, tense, person, gender, kount):
		return [item for item in self.items if item.stem == stem and item.tense == tense
				and item.person == person and item.gender == gender and item.kount == kount]

class Verb(Word):
	def __init__(self, text, yemenitext, translation, root, stem, tense, person, gender, kount):
		super().__init__(text, yemenitext, translation)
		self.root = root
		self.stem = stem
		self.tense = tense
		self.person = person
		self.gender = gender
		self.kount = kount

	@property
	def morphology(self):
		return ", ".join([v for v in [self.root, self.stem.value,
				"גוף %s"%self.person.value if self.person else "",
				self.gender.value if self.gender else "",
				self.kount.value if self.kount else "",
				self.tense.value if self.tense else ""] if v])

class InflectedVerb(Verb):
	def __init__(self, text, yemenitext, translation, root, stem, tense, person, gender, kount, operson, ogender, okount):
		super().__init__(text, yemenitext, translation, root, stem, tense, person, gender, kount)
		self.operson = operson
		self.ogender = ogender
		self.okount = okount

	@property
	def morphology(self):
		a = ", ".join([v for v in [
				self.root, self.stem.value,
				"גוף %s"%self.person.value if self.person else "",
				self.gender.value if self.gender else "",
				self.kount.value if self.kount else "",
				self.tense.value if self.tense else ""] if v])
		b = ", ".join([v for v in [
				"גוף %s"%self.operson.value,
				self.ogender.value if self.ogender else "",
				self.okount.value] if v])
		return a + " + " + b

def parse(text):
	if not text:
		return []
#	text = text.replace("[", "").replace("]", "")
#	text = re.sub("\{[^}]+\}", "", text)
#	text = re.sub("\u05b0\u05b0", "\u05b0", text)
#	text = re.sub("\u05bc\u05bc", "\u05bc", text)
	values = []
	for value in text.split(" "):
		if "|" in value:
			text, yemenitext = value.split("|")
		else:
			text = value
			yemenitext = None
		if not text:
			continue
		if text.endswith(")"):
			l = text[-2]
			text = text[:-3]
			values.append((text, yemenitext))
			values.append((text + l, yemenitext))
		else:
			values.append((text, yemenitext))
	return values

def cellvalue(cell):
	if isinstance(cell, MergedCell):
		for ranje in cell.parent.merged_cells.ranges:
			if cell.coordinate in ranje:
				t = cell.parent.cell(ranje.top[0][0], ranje.top[0][1]).value
			else:
				t = cell.value

class Aramaic:
	def __init__(self):
		print ("aramaic.init")
		self.names = []
		wbook = openpyxl.load_workbook(filename="db/nouns.xlsx", read_only=False)

		print ("PERSONS")
		self.persons = []
		for cols in wbook["persons"].iter_rows(values_only=False):
			gender = Gender[cols[1].value.upper()]
			for text, yemenitext in parse(cols[0].value):
				self.persons.append(PersonName(text, yemenitext, gender))
		print ("PLACES")
		self.places = []
		for cols in wbook["places"].iter_rows(values_only=False):
			for text, yemenitext in parse(cols[0].value):
				self.places.append(PlaceName(text, yemenitext))
		self.names = self.persons + self.places

		print ("ADVERBS")
		self.adverbs = []
		sheet = wbook["inflected adverbs"]
		rows = list(sheet.iter_rows(values_only=True))
		for cols in sheet.iter_rows(min_row=5, values_only=True):
			# ADVERBS
			translation = cols[1]
			for text, yemenitext in parse(cols[0]):
				self.adverbs.append(Adverb(text, yemenitext, translation))
			# INFLECTIONS
			for i in [2, 4, 6, 8, 10, 12, 14, 16, 18, 20]:
				p = rows[0][i]; person = Person["P%s"%p]
				g = (rows[1][i] or "").upper(); gender = Gender[g] if g else None
				c = rows[2][i].upper(); kount = Count[c]
				translation = cols[i + 1]
				for text, yemenitext in parse(cols[i]):
					self.adverbs.append(InflectedAdverb(text, yemenitext, translation, person, gender, kount))
		sheet = wbook["adverbs"]
		for cols in sheet.iter_rows(values_only=True):
			translation = cols[2]
			for text, yemenitext in parse(cols[1]):
				self.adverbs.append(Adverb(text, yemenitext, translation))

		print ("PRONOUNS")
		self.pronouns = []
		for cols in wbook["pronouns"].iter_rows(values_only=True):
			p = cols[0]; person = Person["P%s"%p] if p else None
			g = (cols[1] or "").upper(); gender = Gender[g] if g else None
			c = (cols[2] or "").upper(); kount = Count[c] if c else None
			translation = cols[4]
			for text, yemenitext in parse(cols[3]):
				self.pronouns.append(Pronoun(text, yemenitext, translation, person, gender, kount))

		print ("NOUNS")
		self.nouns = []
		rows = list(wbook["nouns"].iter_rows(values_only=True))
		for cols in wbook["nouns"].iter_rows(min_row=5, min_col=1, max_col=54, values_only=True):
			# SINGULAR
			gender = Gender[cols[0].upper()]
			for i in [1, 3, 5]:
				translation = cols[i + 1]
				for text, yemenitext in parse(cols[i]):
					self.nouns.append(Noun(text, yemenitext, translation, gender, Count.S))
			# SINGULAR + OBJECT
			for colno in [8, 10, 12, 14, 16, 18, 20, 24, 26]:
				p = rows[0][colno - 1]; operson = Person["P%s"%p] if p else None
				g = (rows[1][colno - 1] or "").upper(); ogender = Gender[g] if g else None
				c = (rows[2][colno - 1] or "").upper(); okount = Count[c] if c else None
				translation = cols[colno]
				for text, yemenitext in parse(cols[colno - 1]):
					self.nouns.append(InflectedNoun(text, yemenitext, translation, gender, Count.S, operson, ogender, okount))
			# PLURAL
			for i in [27, 29, 31]:
				translation = cols[i + 1]
				for text, yemenitext in parse(cols[i]):
					translation = cols[i + 1]
					self.nouns.append(Noun(text, yemenitext, translation, gender, Count.P))
			# PLURAL + OBJECT
			for colno in [34, 36, 38, 40, 42, 44, 46, 48, 50, 52]:
				p = rows[0][colno - 1]; operson = Person["P%s"%p] if p else None
				g = (rows[1][colno - 1] or "").upper(); ogender = Gender[g] if g else None
				c = (rows[2][colno - 1] or "").upper(); okount = Count[c] if c else None
				translation = cols[colno]
				for text, yemenitext in parse(cols[colno - 1]):
					self.nouns.append(InflectedNoun(text, yemenitext, translation, gender, Count.P, operson, ogender, okount))

		print ("CARDINALS")
		self.cardinals = []
		for cols in wbook["cardinals"].iter_rows(min_row=2, max_row=21, values_only=True):
			value = int(cols[0])
			translation = cols[2]
			for text, yemenitext in parse(cols[1]):
				self.cardinals.append(Cardinal(text, yemenitext, translation, value, Gender.M))
			translation = cols[4]
			for text, yemenitext in parse(cols[3]):
				self.cardinals.append(Cardinal(text, yemenitext, translation, value, Gender.F))
		print ("ORDINALS")
		self.ordinals = []
		for cols in wbook["ordinals"].iter_rows(min_row=2, values_only=True):
			value = int(cols[0])
			translation = cols[2]
			for text, yemenitext in parse(cols[1]):
				self.ordinals.append(Ordinal(text, yemenitext, translation, value, Gender.M))
			translation = cols[4]
			for text, yemenitext in parse(cols[3]):
				self.ordinals.append(Ordinal(text, yemenitext, translation, value, Gender.F))
		self.numerals = self.cardinals + self.ordinals

		print ("VERBS")
		self.verbs = []
		self.verbtables = {}
		wbook = openpyxl.load_workbook(filename="db/verbs.xlsx")
		roots = ["אתי", "אמר", "יהב", "יתב", "קום", "עלל"]
		for root in roots:
			verbtable = VerbTable(root)
			sheet = wbook[root]
			rows = list(sheet.iter_rows(values_only=True))
			for cols in sheet.iter_rows(min_row=7):
				cell = cols[0]
				t = None
				if isinstance(cell, MergedCell):
					for ranje in sheet.merged_cells.ranges:
						if cell.coordinate in ranje:
							t = sheet.cell(ranje.top[0][0], ranje.top[0][1]).value
				else:
					t = cell.value
				tense = Tense(t)# if t else None
				p = cols[1].value;
				person = Person["P%s"%p] if p else None
				g = (cols[2].value or "").upper();
				gender = Gender[g] if g else None
				c = (cols[3].value or "").upper();
				kount = Count[c] if c else None

				# PEIL
				translation = cols[6].value
				for text, yemenitext in parse(cols[5].value):
					verb = Verb(text, yemenitext, translation, root, Stem.PEIL, tense, person, gender, kount)
					self.verbs.append(verb)
					verbtable.items.append(verb)
				# PEIL + OBJECT
				for colno in [8, 10, 12, 14, 16, 18, 20, 22, 24, 26]:
					p = rows[2][colno - 1]; operson = Person["P%s"%p] if p else None
					g = (rows[3][colno - 1] or "").upper(); ogender = Gender[g] if g else None
					c = (rows[4][colno - 1] or "").upper(); okount = Count[c] if c else None
					translation = cols[colno].value
					for text, yemenitext in parse(cols[colno - 1].value):
						verb = InflectedVerb(text, yemenitext, translation, root, Stem.PEIL, tense, person, gender, kount, operson, ogender, okount)
						self.verbs.append(verb)
						verbtable.items.append(verb)
				# PAEL
				translation = cols[28].value
				for text, yemenitext in parse(cols[27].value):
					verb = Verb(text, yemenitext, translation, root, Stem.PAEL, tense, person, gender, kount)
					self.verbs.append(verb)
					verbtable.items.append(verb)
				# PAEL + OBJECT
				for colno in [30, 32, 34, 36, 38, 40, 42, 44, 46, 48]:
					p = rows[2][colno - 1]; operson = Person["P%s"%p] if p else None
					g = (rows[3][colno - 1] or "").upper(); ogender = Gender[g] if g else None
					c = (rows[4][colno - 1] or "").upper(); okount = Count[c] if c else None
					translation = cols[colno - 1].value
					for text, yemenitext in parse(cols[colno - 1].value):
						verb = InflectedVerb(text, yemenitext, translation, root, Stem.PAEL, tense, person, gender, kount, operson, ogender, okount)
						self.verbs.append(verb)
						verbtable.items.append(verb)
				# AFEL
				translation = cols[50].value
				for text, yemenitext in parse(cols[49].value):
					verb = Verb(text, yemenitext, translation, root, Stem.AFEL, tense, person, gender, kount)
					self.verbs.append(verb)
					verbtable.items.append(verb)
				# AFEL + OBJECT
				for colno in [52, 54, 56, 58, 60, 62, 64, 66, 68, 70]:
					p = rows[2][colno - 1]; operson = Person["P%s"%p] if p else None
					g = (rows[3][colno - 1] or "").upper(); ogender = Gender[g] if g else None
					c = (rows[4][colno - 1] or "").upper(); okount = Count[c] if c else None
					translation = cols[colno - 1].value
					for text, yemenitext in parse(cols[colno - 1].value):
						verb = InflectedVerb(text, yemenitext, translation, root, Stem.AFEL, tense, person, gender, kount, operson, ogender, okount)
						self.verbs.append(verb)
						verbtable.items.append(verb)

			self.verbtables[root] = verbtable
		print ("aramaic.init end")

	@property
	def words(self):
		l = self.adverbs + self.pronouns + self.nouns + self.verbs + self.numerals + self.names
		l.sort(key=lambda x:strip_hebrew_points(x.text))
		return l

	def search(self, text):
		return [word for word in self.words if strip_hebrew_double_points(word.fulltext) == text]

	def baresearch(self, text):
		text = strip_hebrew_points(text)
		return [word for word in self.words if strip_hebrew_points(word.fulltext) == text]

if __name__ == "__main__":
	a = Aramaic()
